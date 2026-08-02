from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from openpyxl import load_workbook

from members.importing import (
    ImportValidationError,
    WorkbookMemberImporter,
    export_membership_backup,
)


class Command(BaseCommand):
    help = "Safely import member memberships from an XLSX/XLSM workbook."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to the .xlsx or .xlsm workbook.")
        parser.add_argument("--sheet", help="Sheet name. Defaults to the first sheet.")
        parser.add_argument(
            "--level",
            choices=["district", "municipality", "rural_municipality", "ward", "unknown", "other"],
            help="Organization level. Inferred from the workbook title when omitted.",
        )
        parser.add_argument("--unit-name", help="Organization unit/registry name.")
        parser.add_argument("--dry-run", action="store_true", help="Validate and simulate without changing people or memberships.")
        parser.add_argument("--update-existing", action="store_true", help="Apply changed workbook values to matching memberships.")
        parser.add_argument("--report", help="CSV report path. Recommended for every production import.")
        parser.add_argument("--backup", help="JSON backup path. Generated automatically for real imports when omitted.")
        parser.add_argument("--list-sheets", action="store_true", help="List sheet names and exit.")

    def handle(self, *args, **options):
        path = Path(options["workbook"]).expanduser()
        if not path.is_absolute():
            path = Path(settings.BASE_DIR) / path

        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        if options["list_sheets"]:
            workbook = load_workbook(path, read_only=True, data_only=True)
            self.stdout.write("Available sheets:")
            for sheet in workbook.sheetnames:
                self.stdout.write(f"- {sheet}")
            return

        timestamp = timezone.now().strftime("%Y%m%d-%H%M%S")
        report_path = options["report"]
        if not report_path:
            report_path = Path(settings.BASE_DIR) / "import_reports" / f"{path.stem}-{timestamp}.csv"

        if not options["dry_run"]:
            backup_path = options["backup"] or (
                Path(settings.BASE_DIR) / "backups" / f"members-before-{path.stem}-{timestamp}.json"
            )
            backup = export_membership_backup(backup_path)
            self.stdout.write(self.style.WARNING(f"Created pre-import backup: {backup}"))

        importer = WorkbookMemberImporter(
            path,
            sheet_name=options["sheet"],
            level=options["level"],
            unit_name=options["unit_name"],
            update_existing=options["update_existing"],
            dry_run=options["dry_run"],
            report_path=report_path,
        )

        try:
            batch, summary, outcomes = importer.run()
        except ImportValidationError as exc:
            raise CommandError(str(exc)) from exc

        style = self.style.WARNING if options["dry_run"] else self.style.SUCCESS
        self.stdout.write(style("DRY RUN — no membership changes were committed." if options["dry_run"] else "Import completed."))
        self.stdout.write(f"Import batch: {batch.public_id}")
        self.stdout.write(f"Report: {report_path}")
        self.stdout.write("")
        self.stdout.write("Reconciliation")
        labels = [
            ("Total workbook rows", summary.total_workbook_rows),
            ("Header rows", summary.header_rows),
            ("Blank rows", summary.blank_rows),
            ("Non-data/summary rows", summary.non_data_rows),
            ("Valid data rows", summary.valid_data_rows),
            ("Imported people", summary.imported_people),
            ("Imported memberships", summary.imported_memberships),
            ("Updated people", summary.updated_people),
            ("Updated memberships", summary.updated_memberships),
            ("Unchanged records", summary.unchanged_records),
            ("Multiple-membership people", summary.multiple_membership_people),
            ("Potential duplicates", summary.potential_duplicates),
            ("Warning rows", summary.warning_rows),
            ("Failed rows", summary.failed_rows),
            ("Database people before", summary.people_before),
            ("Database people after", summary.people_after),
            ("Database memberships before", summary.memberships_before),
            ("Database memberships after", summary.memberships_after),
            ("Records deleted", summary.records_deleted),
        ]
        for label, value in labels:
            self.stdout.write(f"{label}: {value}")

        failed = [outcome for outcome in outcomes if outcome.status == "failed"]
        if failed:
            self.stdout.write(self.style.ERROR("Failed rows:"))
            for outcome in failed:
                self.stdout.write(f"- Row {outcome.row_number}: {outcome.error}")
            raise CommandError(f"Import completed with {len(failed)} failed row(s). Review the CSV report.")
