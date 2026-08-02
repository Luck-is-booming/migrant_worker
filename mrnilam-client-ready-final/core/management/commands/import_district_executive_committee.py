import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import TeamMember


DESIGNATIONS = {
    "अध्यक्ष": "Chairperson",
    "उपाध्यक्ष": "Vice-chairperson",
    "सचिव": "Secretary",
    "सह-सचिव": "Joint Secretary",
    "कोषाध्यक्ष": "Treasurer",
    "सदस्य": "Member",
}


class Command(BaseCommand):
    help = (
        "Import the district executive committee workbook into bilingual Team Member "
        "records. Existing nonblank admin content is not overwritten."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="data/Mrn District Level exaucative commetee.xlsm",
        )
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--report", default="import_reports/district-committee.csv")

    def handle(self, *args, **options):
        source = Path(options["file"]).expanduser().resolve()
        if not source.exists():
            raise CommandError(f"Committee workbook not found: {source}")
        report = Path(options["report"]).expanduser().resolve()
        report.parent.mkdir(parents=True, exist_ok=True)

        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        outcomes = []
        created = updated = skipped = 0

        with transaction.atomic():
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=3, values_only=True), start=3
            ):
                serial, name_ne, address_ne, designation_ne, phone = values[:5]
                name_ne = str(name_ne or "").strip()
                if not name_ne:
                    skipped += 1
                    outcomes.append((row_number, "skipped", "Blank name", ""))
                    continue
                designation_ne = str(designation_ne or "").strip()
                defaults = {
                    "designation_ne": designation_ne or "सदस्य",
                    "designation_en": DESIGNATIONS.get(designation_ne, ""),
                    "address_ne": str(address_ne or "").strip(),
                    "phone": str(phone or "").split(".", 1)[0].strip(),
                    "sort_order": int(serial or 0),
                    "is_active": True,
                }
                member, was_created = TeamMember.objects.get_or_create(
                    name_ne=name_ne,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                    status = "created"
                else:
                    changes = []
                    for field, value in defaults.items():
                        current = getattr(member, field)
                        if value and not current:
                            setattr(member, field, value)
                            changes.append(field)
                    if changes:
                        member.save(update_fields=changes)
                        updated += 1
                        status = "updated_blank_fields"
                    else:
                        skipped += 1
                        status = "unchanged"
                outcomes.append(
                    (
                        row_number,
                        status,
                        "",
                        "Review English spelling and address translation in admin.",
                    )
                )

            if options["dry_run"]:
                transaction.set_rollback(True)

        with report.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["source_row", "status", "error", "review_note"])
            writer.writerows(outcomes)

        prefix = "DRY RUN: " if options["dry_run"] else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}committee rows processed: {created} created, "
                f"{updated} updated, {skipped} unchanged/skipped."
            )
        )
        self.stdout.write(f"Report: {report}")
        self.stdout.write(
            self.style.WARNING(
                "The source contains Nepali names and addresses only. Review every English "
                "name/address field in admin before describing the English page as complete."
            )
        )
