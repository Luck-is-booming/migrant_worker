"""
Reset and import MRN members into the members.Member table without merging district
and municipality registry records.

Expected files beside manage.py:
- Life Time Member of MRN District(1).xlsm
- MRN Ilam nagar level all Member List(1).xlsm

Run:
    pip install openpyxl
    python3 manage.py shell < reset_import_members_unmerged.py

Important:
- This deletes ONLY members.Member records.
- It does NOT delete admin users, organization info, team members, payments, or blog posts.
"""

from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from members.models import Member

BASE_DIR = Path(settings.BASE_DIR)
DISTRICT_FILE = BASE_DIR / "Life Time Member of MRN District.xlsm"
ILAM_FILE = BASE_DIR / "MRN Ilam nagar level all Member List.xlsm"

DISTRICT_UNIT = "Ilam District"
ILAM_UNIT = "Ilam Municipality"


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def clean_phone(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        value = int(value)
    return str(value).strip()


def clean_member_no(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        value = int(value)
    return str(value).strip()


def number_to_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def get_membership_type(value):
    text = clean_text(value)
    if "आजिवन" in text or "आजीवन" in text:
        return "life"
    return "general"


def get_status(value):
    text = clean_text(value).lower()
    if text in ("active", "सक्रिय"):
        return "active"
    if text in ("expire", "expired", "निष्क्रिय"):
        return "expired"
    return "unknown"


def detect_municipality(address):
    text = clean_text(address)

    if "इ.न.पा" in text or "इलाम" in text:
        return "Ilam Municipality"
    if "देउमाइ" in text or "देउमाई" in text:
        return "Deumai Municipality"
    if "सूर्योदय" in text or "सुर्योदय" in text:
        return "Suryodaya Municipality"
    if "सन्दकपुर" in text or "संदकपुर" in text:
        return "Sandakpur Rural Municipality"
    if "फाकफोकथुम" in text or "फाकफोकथुम" in text:
        return "Phakphokthum Rural Municipality"
    if "माङ्सेवुङ" in text or "माङ्सेवुङ्" in text or "माङसेबुङ" in text:
        return "Mangsebung Rural Municipality"

    return ""


def make_member(*, level, unit_name, name, address, designation, member_no, member_type, country, phone, status, sort_order, municipality):
    member_no = clean_member_no(member_no)
    name = clean_text(name)

    if not name or not member_no:
        return None

    return Member(
        name_ne=name,
        name_en="",
        membership_number=member_no,
        membership_number_int=number_to_int(member_no),
        membership_type=get_membership_type(member_type),
        status=status,
        level=level,
        unit_name=unit_name,
        municipality=municipality,
        address=clean_text(address),
        designation=clean_text(designation),
        destination_country=clean_text(country),
        phone=clean_phone(phone),
        show_phone_publicly=False,
        is_public=True,
        sort_order=sort_order,
    )


def import_district_members():
    wb = load_workbook(DISTRICT_FILE, read_only=True, data_only=True)
    ws = wb.active

    members = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        # District columns:
        # s.n, name, address, designation, member_no, date, type, country, phone, fee
        values = list(row) + [None] * 10
        sn, name, address, designation, member_no, _date, member_type, country, phone, _fee = values[:10]

        if clean_text(sn).lower().startswith("total"):
            continue

        obj = make_member(
            level="district",
            unit_name=DISTRICT_UNIT,
            name=name,
            address=address,
            designation=designation,
            member_no=member_no,
            member_type=member_type,
            country=country,
            phone=phone,
            status="active",
            sort_order=number_to_int(member_no) or 0,
            municipality=detect_municipality(address),
        )
        if obj:
            members.append(obj)

    return members


def import_ilam_municipality_members():
    wb = load_workbook(ILAM_FILE, read_only=True, data_only=True)
    ws = wb.active

    members = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        # Ilam municipality columns:
        # s.n, name, address, date, designation, member_no, type, country, phone, fee, remarks
        values = list(row) + [None] * 11
        sn, name, address, _date, designation, member_no, member_type, country, phone, _fee, remarks = values[:11]

        if clean_text(sn).lower().startswith("total"):
            continue

        obj = make_member(
            level="municipality",
            unit_name=ILAM_UNIT,
            name=name,
            address=address,
            designation=designation,
            member_no=member_no,
            member_type=member_type,
            country=country,
            phone=phone,
            status=get_status(remarks),
            sort_order=number_to_int(sn) or 0,
            municipality=ILAM_UNIT,
        )
        if obj:
            members.append(obj)

    return members


for file_path in [DISTRICT_FILE, ILAM_FILE]:
    if not file_path.exists():
        raise FileNotFoundError(f"Missing Excel file beside manage.py: {file_path.name}")

with transaction.atomic():
    old_count = Member.objects.count()
    Member.objects.all().delete()

    district_members = import_district_members()
    ilam_members = import_ilam_municipality_members()

    Member.objects.bulk_create(district_members + ilam_members)

print(f"Deleted old members.Member rows: {old_count}")
print(f"Imported district registry rows: {len(district_members)}")
print(f"Imported Ilam municipality registry rows: {len(ilam_members)}")
print(f"Total new unmerged member rows: {len(district_members) + len(ilam_members)}")
print("Example result: Nabin can exist as District #4 and Ilam Municipality #1 separately.")
