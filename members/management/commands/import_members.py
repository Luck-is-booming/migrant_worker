import re
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from members.models import Member
from members.name_utils import romanize_nepali_name

def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_phone(value):
    if value is None:
        return ""

    if isinstance(value, float):
        value = int(value)

    return str(value).strip()


def get_import_member_number(membership_number, sn, name):
    """
    Used for update_or_create lookup.

    Main idea:
    - If Excel has real member number, use it.
    - If member number is blank, use SN as fallback.
    - This prevents many blank member numbers from becoming one duplicate record.
    """
    membership_number = clean_value(membership_number)
    sn = clean_value(sn)
    name = clean_value(name)

    if membership_number:
        return membership_number

    if sn:
        return f"SN-{sn}"

    return f"NAME-{name}"


def get_membership_type(value):
    value = clean_value(value)

    if "आजिवन" in value:
        return "life"

    if "साधारण" in value:
        return "general"

    return "general"


def get_status(value, default="active"):
    value = clean_value(value).lower()

    if value == "active":
        return "active"

    if value in ["expire", "expired"]:
        return "expired"

    return default


def extract_ward_no(address):
    address = clean_value(address)

    match = re.search(r"(\d+)", address)
    if match:
        return int(match.group(1))

    return None


def is_valid_member_row(sn, name):
    name = clean_value(name)

    if not name:
        return False

    if name.startswith("आजिवन सदस्य"):
        return False

    if clean_value(sn).lower().startswith("total"):
        return False

    return True


class Command(BaseCommand):
    help = "Import MRN members from Excel files"

    def add_arguments(self, parser):
        parser.add_argument(
            "--district",
            type=str,
            help="Path to district lifetime member Excel file",
        )
        parser.add_argument(
            "--nagar",
            type=str,
            help="Path to Ilam nagar member Excel file",
        )

    def handle(self, *args, **options):
        total_created = 0
        total_updated = 0

        district_path = options.get("district")
        nagar_path = options.get("nagar")

        if district_path:
            created, updated = self.import_district_file(district_path)
            total_created += created
            total_updated += updated

        if nagar_path:
            created, updated = self.import_nagar_file(nagar_path)
            total_created += created
            total_updated += updated

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Created: {total_created}, Updated: {total_updated}"
            )
        )

    def import_district_file(self, file_path):
        file_path = Path(file_path)

        if not file_path.exists():
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return 0, 0

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        created_count = 0
        updated_count = 0
        blank_rows = 0

        # District file:
        # SN = col 0
        # Name = col 1
        # Address = col 2
        # Designation = col 3
        # Membership number = col 4
        # Membership type = col 6
        # Destination country = col 7
        # Phone = col 8
        for row in sheet.iter_rows(min_row=4, max_col=9, values_only=True):
            sn = row[0]
            name = row[1]

            if not clean_value(name):
                blank_rows += 1
                if blank_rows >= 20:
                    break
                continue

            blank_rows = 0

            address = row[2]
            designation = row[3]
            membership_number = row[4]
            membership_type = row[6]
            destination_country = row[7]
            phone = row[8]

            if not is_valid_member_row(sn, name):
                continue

            import_member_number = get_import_member_number(
                membership_number=membership_number,
                sn=sn,
                name=name,
            )

            obj, created = Member.objects.update_or_create(
                level="district",
                unit_name="Ilam District",
                membership_number=import_member_number,
                defaults={
                    "name_ne": clean_value(name),
                    "name_en": romanize_nepali_name(name),
                    "address": clean_value(address),
                    "designation": clean_value(designation),
                    "membership_type": get_membership_type(membership_type),
                    "status": "active",
                    "municipality": "",
                    "ward_no": extract_ward_no(address),
                    "destination_country": clean_value(destination_country),
                    "phone": clean_phone(phone),
                    "show_phone_publicly": False,
                    "is_public": True,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"District file imported. Created: {created_count}, Updated: {updated_count}"
            )
        )

        return created_count, updated_count

    def import_nagar_file(self, file_path):
        file_path = Path(file_path)

        if not file_path.exists():
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return 0, 0

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        created_count = 0
        updated_count = 0
        blank_rows = 0

        # Nagar / municipality file:
        # SN = col 0
        # Name = col 1
        # Address = col 2
        # Designation = col 4
        # Membership number = col 5
        # Membership type = col 6
        # Destination country = col 7
        # Phone = col 8
        # Status = col 10
        for row in sheet.iter_rows(min_row=4, max_col=11, values_only=True):
            sn = row[0]
            name = row[1]

            if not clean_value(name):
                blank_rows += 1
                if blank_rows >= 20:
                    break
                continue

            blank_rows = 0

            address = row[2]
            designation = row[4]
            membership_number = row[5]
            membership_type = row[6]
            destination_country = row[7]
            phone = row[8]
            status = row[10]

            if not is_valid_member_row(sn, name):
                continue

            import_member_number = get_import_member_number(
                membership_number=membership_number,
                sn=sn,
                name=name,
            )

            obj, created = Member.objects.update_or_create(
                level="municipality",
                unit_name="Ilam Municipality",
                membership_number=import_member_number,
                defaults={
                    "name_ne": clean_value(name),
                    "name_en": romanize_nepali_name(name),
                    "address": clean_value(address),
                    "designation": clean_value(designation),
                    "membership_type": get_membership_type(membership_type),
                    "status": get_status(status, default="active"),
                    "municipality": "Ilam Municipality",
                    "ward_no": extract_ward_no(address),
                    "destination_country": clean_value(destination_country),
                    "phone": clean_phone(phone),
                    "show_phone_publicly": False,
                    "is_public": True,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

            total_processed = created_count + updated_count
            if total_processed % 50 == 0:
                self.stdout.write(f"Processed {total_processed} municipality members...")

        self.stdout.write(
            self.style.SUCCESS(
                f"Nagar file imported. Created: {created_count}, Updated: {updated_count}"
            )
        )

        return created_count, updated_count